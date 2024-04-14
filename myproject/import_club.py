import os
import sys
import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "myproject.settings")
sys.path.append(os.path.dirname(os.path.abspath(__file__)))


django.setup()

from openpyxl import load_workbook
from boardgames.models import Club

def import_clubs_from_excel(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Assuming the first row is the header, start reading from the second row
    for row in sheet.iter_rows(min_row=2, values_only=True):
        club_name, club_link, contact, address, city = row
        club = Club(name=club_name, link=club_link, contact=contact, address=address, city=city)
        club.save()

if __name__ == "__main__":
    # Replace 'your_excel_file.xlsx' with the path to your Excel file
    import_clubs_from_excel('Clubs.xlsx')
